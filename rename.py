from openpyxl import load_workbook
import os


wb = load_workbook("data.xlsx")  # Work Book
ws = wb.get_sheet_by_name('photography teachers')  # Work Sheet
column = ws['A']  # Column
names = [column[x].value for x in range(len(column))]


os.chdir("Photography Teachers Participation Certificates")

par_counter = 1
for name in names:
    src = "Slide" + str(par_counter) + ".png"
    des = name + ".png"
    par_counter += 1

    os.rename(src,des)