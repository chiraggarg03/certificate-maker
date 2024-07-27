import win32com.client as comclt
import time
import keyboard
from openpyxl import load_workbook
import os


def finrep(fin, rep):
    keyboard.press_and_release("control+f")
    keyboard.write(fin)
    keyboard.press_and_release("enter")
    keyboard.press_and_release("tab")
    keyboard.press_and_release("enter")
    keyboard.write(rep)


wb = load_workbook("data.xlsx")  # Work Book
ws = wb.get_sheet_by_name('film jury')  # Work Sheet
column = ws['B']  # Column
names = [column[x].value for x in range(len(column))]
column2 = ws['C']
schools = [column2[x].value for x in range(len(column2))]


hashname = "9BC0CB2ED6DEF7AFEA9395AB78790A0C"
hashschool = "32C19AA31D5ACE768BFF6E89E3EFB44C163D750D"


wsh = comclt.Dispatch("WScript.Shell")
wsh.AppActivate("PowerPoint")  # select another application

for i in range(len(names)):
    keyboard.press_and_release("control+shift+d")

    keyboard.press_and_release("up")

    finrep(hashname, names[i])

    finrep(hashschool, schools[i])

    keyboard.press_and_release("esc")
    keyboard.press_and_release("esc")
    keyboard.press_and_release("down")
    time.sleep(1.3)

'''
keyboard.press_and_release("control+shift+s")
keyboard.write("Certs")
keyboard.press_and_release("tab")
keyboard.write("gp")
keyboard.press_and_release("Enter")
time.sleep(1.3)
keyboard.press_and_release("Enter")


time.sleep(10)

keyboard.press_and_release("Enter")
time.sleep(2)

os.chdir("Certs")

par_counter = 1
for name in names:
    src = "Slide" + str(par_counter) + ".png"
    des = str(par_counter) + " - " + name + ".png"
    par_counter += 1

    os.rename(src,des)

    
    

'''
